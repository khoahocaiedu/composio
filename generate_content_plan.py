import os
import sys
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing openpyxl...")
    # Use current python executable from .venv to install openpyxl
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

from datetime import datetime, timedelta

def create_excel_schedule():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Post Schedule"
    
    # Enable grid lines
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Scheduled Date",
        "Message",
        "Image Prompt",
        "Article Link",
        "Status",
        "Post ID",
        "Published Time"
    ]
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        # Style header
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Start date is today
    start_date = datetime.now()
    
    # 30-day content calendar data
    posts_data = [
        # Day 1
        {
            "offset": 0,
            "message": (
                "🎯 LỘ TRÌNH HỌC AI CHO NGƯỜI MỚI BẮT ĐẦU: BẬT PHÁT THÀNH THẠO! 🎯\n\n"
                "Bạn muốn bắt nhịp với làn sóng công nghệ Trí tuệ Nhân tạo nhưng chưa biết bắt đầu từ đâu? Đừng lo lắng! "
                "Trung tâm Đào tạo AI Sao Việt đã xây dựng lộ trình học AI thực chiến tối ưu nhất, đi từ cơ bản đến nâng cao:\n"
                "1. Làm quen với tư duy AI và Prompt Engineering cơ bản.\n"
                "2. Sử dụng thành thạo ChatGPT, Gemini và Claude AI hỗ trợ công việc văn phòng hàng ngày.\n"
                "3. Ứng dụng AI vào thiết kế hình ảnh (Midjourney, Canva AI) và tạo slide tự động.\n"
                "4. Tự động hóa quy trình với Make.com và n8n.\n\n"
                "👉 Đọc bài viết chi tiết để nhận ngay lộ trình miễn phí tại:\n"
                "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/\n\n"
                "---\n"
                "#KhoaHocAI #SaoViet #LoTrinhHocAI #ChatGPT #GoogleGemini #ClaudeAI #AiForBeginners"
            ),
            "image_prompt": (
                "A clean, modern infographic illustration representing an 'AI Learning Roadmap'. "
                "A winding glowing path with stepping stones labeled: 'Basics', 'ChatGPT & Gemini', 'AI Tools', 'Automation'. "
                "At the end of the path is a bright shining brain icon representing mastery. "
                "Futuristic classroom aesthetic, bright blue and white color palette, professional design, highly detailed."
            ),
            "link": "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/"
        },
        # Day 2
        {
            "offset": 1,
            "message": (
                "🚀 GOOGLE GEMINI 3.5 FLASH: KỶ NGUYÊN MỚI CỦA TÁC NHÂN TỰ ĐỘNG (AGENTIC AI)! 🚀\n\n"
                "Google đã trình làng thế hệ Gemini 3.5 Flash với những bước tiến vượt trội về hiệu năng và tốc độ xử lý:\n"
                "⚡ Tốc độ xuất token nhanh gấp 4 lần so với các dòng mô hình Frontier hiện nay.\n"
                "⚡ Cửa sổ ngữ cảnh khổng lồ 1 triệu token cho phép xử lý dữ liệu và văn bản khổng lồ trong chớp mắt.\n"
                "⚡ Thiết kế chuyên biệt cho Agentic AI: tối ưu hóa tư duy suy luận dài hạn, xuất dữ liệu cấu trúc (structured outputs), và gọi hàm đa phương tiện (multimodal function calling).\n\n"
                "👉 Tìm hiểu thêm kiến thức cơ bản về trí tuệ nhân tạo và lộ trình ứng dụng thực tế tại:\n"
                "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/\n\n"
                "---\n"
                "#KhoaHocAI #GoogleGemini #Gemini35Flash #AIAgent #AgenticAI #CongNgheMoi #GoogleCloud"
            ),
            "image_prompt": (
                "A futuristic digital illustration representing 'Gemini 3.5 Flash'. "
                "A sleek, glowing double-helix energy pulse in cyan and magenta colors, shooting forward like a bolt of lightning through a digital server network. "
                "Minimalist modern UI elements floating around, symbolizing speed and agentic actions. "
                "High tech, 3D render style, cinematic lighting."
            ),
            "link": "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/"
        },
        # Day 3
        {
            "offset": 2,
            "message": (
                "⚙️ TỰ ĐỘNG HÓA CÔNG VIỆC VỚI MAKE.COM - PHẦN 1: KHỞI ĐẦU DỄ DÀNG ⚙️\n\n"
                "Bạn mệt mỏi với các công việc lặp đi lặp lại hàng ngày như copy-paste dữ liệu, gửi email báo cáo hay quản lý khách hàng? "
                "Đã đến lúc dùng Make.com để giải phóng sức lao động! Phần 1 của chuỗi bài viết hướng dẫn Make.com sẽ giúp bạn:\n"
                "- Hiểu rõ cơ chế hoạt động của Scenario, Trigger và Action.\n"
                "- Kết nối tài khoản và thiết lập luồng công việc tự động đầu tiên của bạn.\n"
                "- Ứng dụng thực tế: Tự động gửi thông báo từ Gmail sang Telegram/Slack.\n\n"
                "👉 Đọc bài viết hướng dẫn chi tiết từng bước tại:\n"
                "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/\n\n"
                "---\n"
                "#KhoaHocAI #SaoViet #MakeDotCom #TuDongHoa #NoCode #ProductivityHack #WorkflowAutomation"
            ),
            "image_prompt": (
                "A creative digital concept representing work automation. "
                "A series of colorful connected puzzle pieces or visual nodes with icons (Gmail, Google Sheets, Telegram) connected by glowing laser lines. "
                "A friendly cartoon robot hands-on assembling the nodes. "
                "Bright, vibrant, clean vector design, professional white background."
            ),
            "link": "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/"
        },
        # Day 4
        {
            "offset": 3,
            "message": (
                "💼 KHÓA HỌC AI TRONG VĂN PHÒNG: X5 HIỆU SUẤT LÀM VIỆC CÙNG SAO VIỆT! 💼\n\n"
                "Khóa học AI Trong Văn Phòng tại Trung tâm Đào tạo AI Sao Việt là chiếc chìa khóa giúp bạn tối ưu hóa công việc văn phòng bằng công nghệ:\n"
                "✔️ Soạn thảo văn bản, viết email, dịch thuật siêu tốc với ChatGPT và Claude.\n"
                "✔️ Phân tích dữ liệu, xử lý bảng tính Excel phức tạp bằng trợ lý Copilot.\n"
                "✔️ Thiết kế nhanh banner, slide thuyết trình bằng Canva AI & Gamma AI.\n\n"
                "Đăng ký ngay hôm nay để không bị bỏ lại phía sau trong kỷ nguyên số!\n\n"
                "👉 Thông tin chi tiết khóa học tại:\n"
                "https://khoahocai.com.vn/khoa-hoc/\n\n"
                "---\n"
                "#KhoaHocAI #AiTrongVanPhong #SaoViet #DaoTaoAI #ChatGPT #MicrosoftCopilot #TangHieuSuat"
            ),
            "image_prompt": (
                "A professional corporate office setting. "
                "An employee happily looking at a laptop screen where holographic charts and smart AI icons (ChatGPT, Excel AI, Canva) are elegantly floating. "
                "Sunny day, clean modern office background, highly professional and premium look."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc/"
        },
        # Day 5
        {
            "offset": 4,
            "message": (
                "🎨 CHATGPT 2.0 & GPT-4o: BƯỚC NGOẶT MỚI TRONG SÁNG TẠO HÌNH ẢNH 🎨\n\n"
                "OpenAI đã ra mắt bản cập nhật ChatGPT mới nâng cấp toàn diện khả năng tạo và chỉnh sửa hình ảnh trực tiếp:\n"
                "✨ Tương tác trực quan: Nhấp chọn một vùng cụ thể trên hình ảnh để chỉnh sửa, thêm hoặc xóa chi tiết bằng văn bản.\n"
                "✨ Hiểu ngữ nghĩa cực kỳ chính xác nhờ sức mạnh của mô hình đa phương tiện GPT-4o.\n"
                "✨ Thiết kế nhanh chóng các banner, concept nghệ thuật hoặc bản vẽ thiết kế chỉ qua vài câu lệnh.\n\n"
                "👉 Tìm hiểu sâu hơn về kiến thức AI và cách ứng dụng thực tiễn trong công việc tại đây:\n"
                "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/\n\n"
                "---\n"
                "#KhoaHocAI #ChatGPT #OpenAI #GPT4o #AIImages #Midjourney #SángTạoẢnhAI #CongNghe2026"
            ),
            "image_prompt": (
                "A digital canvas showing an abstract colorful painting being edited by a virtual glowing brush held by a robot. "
                "The canvas blends realism and digital cyber art. "
                "Holographic UI panels with sliders showing color values. "
                "Creative, artistic, high-definition digital painting style."
            ),
            "link": "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/"
        },
        # Day 6
        {
            "offset": 5,
            "message": (
                "🧠 CLAUDE 3.5 SONNET & TÍNH NĂNG ARTIFACTS: ĐỈNH CAO HỖ TRỢ LẬP TRÌNH VÀ THIẾT KẾ! 🧠\n\n"
                "Claude 3.5 Sonnet từ Anthropic đang khẳng định vị thế dẫn đầu trong cộng đồng công nghệ nhờ khả năng lập trình và suy luận logic đỉnh cao:\n"
                "💻 Tính năng Artifacts: Hiển thị trực tiếp mã nguồn, ứng dụng web mini, sơ đồ SVG và thiết kế ngay trên màn hình trò chuyện tương tác.\n"
                "💻 Lập trình siêu tốc: Cho phép tạo nhanh ứng dụng web đơn giản (React, HTML5) chỉ từ yêu cầu dạng văn bản.\n"
                "💻 Khả năng viết lách tự nhiên, sâu sắc, lập luận chặt chẽ như một chuyên gia thực thụ.\n\n"
                "👉 Xem thêm định hướng nghề nghiệp và việc làm ngành AI tại:\n"
                "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/\n\n"
                "---\n"
                "#KhoaHocAI #ClaudeAI #Claude35Sonnet #Artifacts #Anthropic #LậpTrìnhAI #AIForDevelopers"
            ),
            "image_prompt": (
                "A split-screen illustration showing Claude AI interface. "
                "On the left, clean code lines being written. "
                "On the right, a dynamic interactive web dashboard (hologram) popping out of the screen. "
                "Sleek dark mode interface with neon purple and blue accents, futuristic developer workstation."
            ),
            "link": "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/"
        },
        # Day 7
        {
            "offset": 6,
            "message": (
                "🔗 N8N AUTOMATION: GIẢI PHÁP MÃ NGUỒN MỞ TỰ ĐỘNG HÓA MẠNH MẼ CHO DOANH NGHIỆP 🔗\n\n"
                "Bên cạnh Make.com, n8n.io đang là một trong những nền tảng tự động hóa được yêu thích nhất nhờ tính linh hoạt và khả năng tự lưu trữ (self-hosted):\n"
                "🔸 Tự do kết nối hàng trăm ứng dụng khác nhau mà không lo giới hạn số lượng tác vụ.\n"
                "🔸 Hỗ trợ viết code Javascript/Python trực tiếp trong các Node để xử lý dữ liệu phức tạp.\n"
                "🔸 Tuyệt đối bảo mật thông tin bằng cách chạy local hoặc cài đặt trên máy chủ riêng.\n\n"
                "👉 Tìm hiểu ngay các bước cơ bản tự động hóa quy trình hàng ngày tại:\n"
                "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/\n\n"
                "---\n"
                "#KhoaHocAI #n8n #Automation #SelfHosted #NoCode #LowCode #SmartWorkflow #TuDongHoaDoanhNghiep"
            ),
            "image_prompt": (
                "An abstract visual representation of server backend automation. "
                "Glowing nodes arranged in a logic flow tree, processing digital light pulses. "
                "A prominent cute robotic node at the center acting as the controller. "
                "Deep dark blue background with bright orange and green highlight lines, network database aesthetic."
            ),
            "link": "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/"
        },
        # Day 8
        {
            "offset": 7,
            "message": (
                "🏫 KHÓA HỌC AI CHO GIÁO VIÊN: ĐỔI MỚI PHƯƠNG PHÁP SOẠN GIẢNG VÀ TƯƠNG TÁC 4.0 🏫\n\n"
                "Trí tuệ nhân tạo đang thay đổi cách chúng ta giáo dục. Nhằm đồng hành cùng quý Thầy Cô, Trung tâm Sao Việt ra mắt khóa học AI cho Giáo viên:\n"
                "📚 Soạn giáo án chi tiết và lên ý tưởng giảng dạy sáng tạo cùng ChatGPT & Gemini.\n"
                "📚 Thiết kế bài giảng slide tương tác sinh động chỉ với một click bằng Gamma AI.\n"
                "📚 Tạo bài tập trắc nghiệm và đánh giá học sinh tự động.\n\n"
                "👉 Đăng ký tham gia ngay khóa học thực chiến để dẫn đầu xu thế giáo dục mới tại:\n"
                "https://khoahocai.com.vn/khoa-hoc-ai-cho-giao-vien/\n\n"
                "---\n"
                "#KhoaHocAI #AiChoGiaoVien #SaoViet #GiaoDuc40 #GammaAI #SoanGiaoAn #TrungTamSaoViet"
            ),
            "image_prompt": (
                "A modern classroom with a teacher standing next to a large digital interactive smart screen. "
                "The screen displays colorful interactive educational modules, 3D models of planets, and friendly AI assistants. "
                "Students watching eagerly. Warm, inviting atmosphere, bright and clean style."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc-ai-cho-giao-vien/"
        },
        # Day 9
        {
            "offset": 8,
            "message": (
                "🛸 GOOGLE DEEPMIND ANTIGRAVITY: TRỢ LÝ AI AGENT LẬP TRÌNH THẾ HỆ MỚI 🛸\n\n"
                "Hãy cùng khám phá 'Antigravity' – một thiết kế Agentic AI đột phá từ đội ngũ Advanced Agentic Coding của Google DeepMind:\n"
                "🚀 Khả năng tự lập kế hoạch hành động chi tiết (planning mode) trước khi lập trình hoặc chỉnh sửa hệ thống.\n"
                "🚀 Tự động chạy thử, bắt lỗi và sửa mã nguồn (self-debugging) một cách độc lập.\n"
                "🚀 Kết nối đa nền tảng, quản lý tệp tin và thực thi các quy trình phức tạp mà không cần con người can thiệp liên tục.\n\n"
                "👉 Đọc bài viết giới thiệu về trí tuệ nhân tạo và tiềm năng ứng dụng thực tế tại:\n"
                "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/\n\n"
                "---\n"
                "#KhoaHocAI #Antigravity #GoogleDeepMind #AgenticAI #AdvancedCoding #AIAgent #TuDongHoaCode"
            ),
            "image_prompt": (
                "An artistic representation of an AI coder agent floating in zero gravity. "
                "A holographic interface showing lines of green code wrapping around a glowing central core. "
                "Deep space theme with subtle stars and blue planetary nebula. "
                "Sleek, futuristic, high-tech digital art, epic lighting."
            ),
            "link": "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/"
        },
        # Day 10
        {
            "offset": 9,
            "message": (
                "⚙️ TỰ ĐỘNG HÓA CÔNG VIỆC VỚI MAKE.COM - PHẦN 2: KẾT NỐI NÂNG CAO VỚI ROUTER & FILTER ⚙️\n\n"
                "Ở phần này, chúng ta sẽ cùng đi sâu vào các cấu trúc rẽ nhánh nâng cao trên Make.com giúp hệ thống hoạt động thông minh hơn:\n"
                "⚡ Cách sử dụng Router để điều hướng dữ liệu theo nhiều luồng khác nhau.\n"
                "⚡ Thiết lập Filter để lọc dữ liệu chuẩn xác, loại bỏ thông tin rác.\n"
                "⚡ Thực hành: Tạo luồng tự động cập nhật Lead từ Facebook Lead Ads vào Google Sheets và gửi thông báo riêng biệt cho từng phòng ban.\n\n"
                "👉 Xem ngay hướng dẫn chi tiết tại:\n"
                "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-2/\n\n"
                "---\n"
                "#KhoaHocAI #MakeDotCom #AdvancedAutomation #RouterAndFilter #NoCode #ProductivityHack #Workflow"
            ),
            "image_prompt": (
                "A clean conceptual diagram showing a single glowing stream of data entering a circular prism (the router) and splitting into three distinct colored streams (blue, gold, magenta) flowing through filter gates. "
                "Polished vector design, soft glows, highly educational and clear layout."
            ),
            "link": "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-2/"
        },
        # Day 11
        {
            "offset": 10,
            "message": (
                "🌆 KHÓA HỌC AI THỰC CHIẾN TẠI TP. HCM - NÂNG CẤP KỸ NĂNG SỐ NGAY HÔM NAY! 🌆\n\n"
                "Trung tâm Đào tạo AI Sao Việt khai giảng liên tục các khóa học AI thực chiến tại TP. Hồ Chí Minh:\n"
                "📍 Học trực tiếp 1-1 hoặc theo nhóm nhỏ, đảm bảo chất lượng tiếp thu.\n"
                "📍 Chương trình học ứng dụng thực tế: Không dạy lý thuyết suông, tập trung làm được việc ngay.\n"
                "📍 Dành cho sinh viên, nhân viên văn phòng, người kinh doanh muốn tối ưu năng suất.\n\n"
                "👉 Xem chi tiết lịch khai giảng và học phí khóa học tại:\n"
                "https://khoahocai.com.vn/khoa-hoc/\n\n"
                "---\n"
                "#KhoaHocAI #SaoViet #HocAITPHCM #DaoTaoAITPHCM #ChatGPT #CongNgheSo #HieuSuatCongViec"
            ),
            "image_prompt": (
                "A beautiful modern computer lab in Ho Chi Minh City, overlooking a panoramic window with landmark buildings like Bitexco or Landmark 81 in the background at sunset. "
                "Students practicing AI tools on screens showing code and design programs. "
                "Warm color tones, premium lighting, realistic photo style."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc/"
        },
        # Day 12
        {
            "offset": 11,
            "message": (
                "🦾 OPENCLAW VÀ XU HƯỚNG PHÁT TRIỂN AI AGENTS TRONG TỰ ĐỘNG HÓA 🦾\n\n"
                "OpenCLAW đang nổi lên như một bộ khung mã nguồn mở linh hoạt hỗ trợ xây dựng các trợ lý tác nhân (AI Agents) tự động cho doanh nghiệp:\n"
                "⚙️ Tích hợp sâu với các mô hình ngôn ngữ lớn (LLMs) để đưa ra quyết định thông minh.\n"
                "⚙️ Khả năng tự thực thi các lệnh gọi API, xử lý tệp tin và tìm kiếm web.\n"
                "⚙️ Giúp doanh nghiệp tự phát triển các bot tự vận hành chuỗi cung ứng, chăm sóc khách hàng và kiểm tra chất lượng code.\n\n"
                "👉 Tìm hiểu ngay cơ hội nghề nghiệp trong ngành trí tuệ nhân tạo tương lai:\n"
                "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/\n\n"
                "---\n"
                "#KhoaHocAI #OpenCLAW #AIAgents #AgenticAI #MachineLearning #NoCode #SmartAutomation #ITNews"
            ),
            "image_prompt": (
                "A sleek futuristic robotic hand (metallic chrome and green neon lines) holding a glowing green holographic matrix cube containing code symbols. "
                "Clean minimalist lab background, sharp focus, professional industrial design style."
            ),
            "link": "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/"
        },
        # Day 13
        {
            "offset": 12,
            "message": (
                "💬 NGHỆ THUẬT PROMPT ENGINEERING: LÀM THẾ NÀO ĐỂ AI HIỂU BẠN CHÍNH XÁC? 💬\n\n"
                "Sự khác biệt giữa một câu trả lời AI sơ sài và một kết quả hoàn hảo nằm ở cách bạn viết câu lệnh (Prompt). Hãy áp dụng 3 nguyên tắc vàng sau:\n"
                "1. Role (Vai trò): Chỉ định vai trò cụ thể cho AI (Ví dụ: 'Hãy đóng vai chuyên gia viết bài...').\n"
                "2. Context (Ngữ cảnh): Cung cấp thông tin nền tảng, đối tượng người đọc và mục tiêu bài viết.\n"
                "3. Constraint (Hạn chế): Giới hạn độ dài, ngôn ngữ, định dạng đầu ra (ví dụ: 'tóm tắt dưới dạng bảng').\n\n"
                "👉 Khám phá thêm lộ trình làm chủ AI cho người mới bắt đầu tại:\n"
                "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/\n\n"
                "---\n"
                "#KhoaHocAI #PromptEngineering #ChatGPT #ClaudeAI #SaoViet #KyNangSo #BiQuyetCongViec"
            ),
            "image_prompt": (
                "A stylized illustration representing communication with AI. "
                "A glowing dialogue bubble filled with structured codes and words connecting a human silhouette to a floating glowing crystal brain. "
                "Dark background with blue and violet lights, conceptual digital illustration style."
            ),
            "link": "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/"
        },
        # Day 14
        {
            "offset": 13,
            "message": (
                "⚙️ TỰ ĐỘNG HÓA CÔNG VIỆC VỚI MAKE.COM - PHẦN 3: XỬ LÝ LỖI (ERROR HANDLING) VÀ TỐI ƯU HÓA ⚙️\n\n"
                "Khi xây dựng các luồng tự động hóa, lỗi hệ thống là điều không thể tránh khỏi. Ở phần 3 này, chúng tôi sẽ hướng dẫn bạn:\n"
                "✔️ Cách sử dụng các directive xử lý lỗi: Commit, Resume, Rollback, Ignore.\n"
                "✔️ Thiết lập cảnh báo tự động gửi về Telegram khi một kịch bản bị crash.\n"
                "✔️ Mẹo tối ưu hóa số lượng vận hành (operations) để tiết kiệm chi phí tài khoản.\n\n"
                "👉 Đọc bài viết hướng dẫn chi tiết cuối cùng trong chuỗi bài tự động hóa tại:\n"
                "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-3/\n\n"
                "---\n"
                "#KhoaHocAI #MakeDotCom #ErrorHandling #AutomateEverything #NoCodeDeveloper #WorkflowOptimization"
            ),
            "image_prompt": (
                "A diagnostic screen of an automation software showing a bug icon being fixed and bypassed by a green shield icon. "
                "Sleek interface dashboard, neon red turning into neon green, technology flat vector art style."
            ),
            "link": "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-3/"
        },
        # Day 15
        {
            "offset": 14,
            "message": (
                "🏢 KHÓA HỌC AI THỰC CHIẾN TẠI BÌNH DƯƠNG - NÂNG TẦM NĂNG SUẤT VẬN HÀNH 🏢\n\n"
                "Trung tâm Đào tạo AI Sao Việt mở rộng cơ sở đào tạo thực chiến tại Bình Dương:\n"
                "🎯 Phù hợp cho nhân viên các doanh nghiệp sản xuất, dịch vụ muốn ứng dụng AI tối ưu báo cáo.\n"
                "🎯 Học thực hành trực tiếp trên máy tính với sự hướng dẫn kèm cặp của chuyên gia.\n"
                "🎯 Cung cấp chứng chỉ hoàn thành khóa học uy tín.\n\n"
                "👉 Đăng ký và nhận ưu đãi học phí ngay tại cơ sở Bình Dương:\n"
                "https://khoahocai.com.vn/khoa-hoc-ai-tai-phuong-phu-loi/\n\n"
                "---\n"
                "#KhoaHocAI #SaoViet #HocAIBinhDuong #DaoTaoAIBinhDuong #PhuLoi #BinhDuongNews #Productivity"
            ),
            "image_prompt": (
                "A modern workspace or tech park building in Binh Duong. "
                "Professional young adults sitting around a modern conference table with screens showing data dashboards and business AI workflows. "
                "Bright daylight, clean architectural lines, corporate theme."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc-ai-tai-phuong-phu-loi/"
        },
        # Day 16
        {
            "offset": 15,
            "message": (
                "⚖️ SO SÁNH: CHATGPT VS GEMINI VS CLAUDE - ĐÂU LÀ TRỢ LÝ AI TỐT NHẤT CHO BẠN? ⚖️\n\n"
                "Mỗi mô hình ngôn ngữ lớn (LLM) hiện nay đều sở hữu những thế mạnh độc bản:\n"
                "🤖 ChatGPT: Đa tài, xử lý dữ liệu và tạo hình ảnh đỉnh cao, phù hợp cho mọi tác vụ tổng hợp.\n"
                "🤖 Gemini: Tốc độ xử lý cực nhanh, tích hợp sâu vào hệ sinh thái Google và tra cứu tin tức thời gian thực xuất sắc.\n"
                "🤖 Claude AI: Logic sắc bén, viết văn phong tự nhiên và xử lý khối lượng tài liệu lập trình siêu đỉnh.\n\n"
                "👉 Hãy trang bị kiến thức nền tảng để lựa chọn và làm chủ công cụ phù hợp nhất tại:\n"
                "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/\n\n"
                "---\n"
                "#KhoaHocAI #ChatGPT #GoogleGemini #ClaudeAI #LLMComparison #AiTools #SaoViet"
            ),
            "image_prompt": (
                "A balanced scale comparing three holographic shields representing OpenAI (green), Google Gemini (cyan), and Anthropic Claude (orange). "
                "A sleek metallic background with subtle computational grid lines. "
                "Clean, symbolic, highly detailed 3D render."
            ),
            "link": "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/"
        },
        # Day 17
        {
            "offset": 16,
            "message": (
                "🚜 KHÓA HỌC AI TẠI ĐỒNG NAI: ĐÀO TẠO NGUỒN NHÂN LỰC CHẤT LƯỢNG CAO 🚜\n\n"
                "Đáp ứng nhu cầu chuyển đổi số mạnh mẽ của các doanh nghiệp, Trung tâm Sao Việt triển khai khóa học AI tại Đồng Nai:\n"
                "💼 Giáo trình thực chiến cập nhật các mô hình AI mới nhất của năm 2026.\n"
                "💼 Giúp doanh nghiệp và cá nhân tự động hóa báo cáo, lập kế hoạch kinh doanh và tối ưu marketing.\n"
                "💼 Thời gian học linh hoạt, có lớp cuối tuần tiện lợi.\n\n"
                "👉 Tìm hiểu thêm chi tiết chương trình đào tạo tại:\n"
                "https://khoahocai.com.vn/khoa-hoc/\n\n"
                "---\n"
                "#KhoaHocAI #SaoViet #HocAIDongNai #DaoTaoAIDongNai #DongNaiNews #ChuyenDoiSo #KỹNăngAI"
            ),
            "image_prompt": (
                "A sleek modern training facility inside a tech building in Dong Nai. "
                "Students working on modern workstations, displaying clean data charts. "
                "Bright natural light coming through a large glass window, professional corporate design."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc/"
        },
        # Day 18
        {
            "offset": 17,
            "message": (
                "🤖 AI AGENTS: TƯƠNG LAI CỦA CÔNG NGHỆ TỰ ĐỘNG HÓA THÔNG MINH 🤖\n\n"
                "Kỷ nguyên của các Chatbot thụ động đang dần nhường chỗ cho các AI Agent (Tác nhân AI) tự vận hành:\n"
                "✔️ Khả năng tự nhận thức mục tiêu và chia nhỏ công việc thành nhiều bước lập kế hoạch.\n"
                "✔️ Chủ động sử dụng các công cụ bên ngoài (gọi API, viết mã chạy thử, đọc cơ sở dữ liệu).\n"
                "✔️ Tự động sửa lỗi khi kết quả không như ý muốn mà không cần nhắc nhở.\n\n"
                "👉 Khám phá hướng đi tiềm năng của ngành Trí tuệ Nhân tạo hiện nay tại:\n"
                "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/\n\n"
                "---\n"
                "#KhoaHocAI #AIAgents #AgenticAI #FutureOfWork #AIAutomation #MachineLearning #CongNgheMoi"
            ),
            "image_prompt": (
                "A conceptual visual of an AI Agent. "
                "A translucent white android figure floating, holding a glowing digital network map where nodes represent actions like 'Analyze', 'Generate', 'Deploy'. "
                "Abstract background, dark blue with glowing particles, cyberpunk corporate style."
            ),
            "link": "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/"
        },
        # Day 19
        {
            "offset": 18,
            "message": (
                "📊 TỐI ƯU HÓA QUY TRÌNH HÀNH CHÍNH VÀ NHÂN SỰ BẰNG TRÍ TUỆ NHÂN TẠO 📊\n\n"
                "Bộ phận Hành chính - Nhân sự (HR) có thể tiết kiệm tới 70% thời gian làm việc nhờ các công cụ AI thực chiến:\n"
                "- Viết JD tuyển dụng và soạn thảo email mời phỏng vấn tự động.\n"
                "- Sử dụng AI để chấm điểm, sàng lọc CV ứng viên nhanh chóng.\n"
                "- Tự động hóa bảng lương và báo cáo hành chính thông qua n8n/Make.\n\n"
                "👉 Đăng ký khóa học AI văn phòng thực chiến để tối ưu hóa quy trình làm việc ngay hôm nay:\n"
                "https://khoahocai.com.vn/khoa-hoc-ai-trong-van-van-phong/\n\n"
                "---\n"
                "#KhoaHocAI #AiTrongNhanSu #HRTech #SaoViet #AiTrongVanPhong #TựĐộngHóa #HRAutomation"
            ),
            "image_prompt": (
                "A stylized graphical screen illustrating candidate filtering. "
                "A computer display showing resumes being scanned, with green checkmarks and AI match scores popping up on screen. "
                "Clean office background, vector design, soft focus, professional colors."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc-ai-trong-van-van-phong/"
        },
        # Day 20
        {
            "offset": 19,
            "message": (
                "🌊 KHÓA HỌC AI THỰC CHIẾN TẠI VŨNG TÀU - LÀM CHỦ CÔNG NGHỆ DẪN ĐẦU XU HƯỚNG 🌊\n\n"
                "Bắt kịp xu thế thời đại số với khóa học AI thực chiến tại thành phố Vũng Tàu cùng Sao Việt:\n"
                "💪 Học viên được kèm cặp 1-1, học đến đâu thực hành làm được việc đến đó.\n"
                "💪 Xây dựng tư duy giải quyết vấn đề bằng AI để tăng năng suất làm việc cá nhân.\n"
                "💪 Cơ hội tuyệt vời cho sinh viên và người đi làm tại khu vực Bà Rịa - Vũng Tàu.\n\n"
                "👉 Đăng ký học trực tiếp tại Vũng Tàu qua liên kết sau:\n"
                "https://khoahocai.com.vn/khoa-hoc-ai-tai-phuong-phuoc-long/\n\n"
                "---\n"
                "#KhoaHocAI #SaoViet #HocAIVungTau #DaoTaoAIVungTau #PhuocLong #VungTauNews #KyNangSo"
            ),
            "image_prompt": (
                "A digital technology classroom located in Vung Tau with ocean scenery visible through a clean window. "
                "Students working on dual screens showing data analytics and neural networks. "
                "Beautiful morning light, photorealistic style, premium look."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc-ai-tai-phuong-phuoc-long/"
        },
        # Day 21
        {
            "offset": 20,
            "message": (
                "📈 XỬ LÝ BÁO CÁO TÀI CHÍNH VÀ DỮ LIỆU EXCEL NHANH GẤP 5 LẦN VỚI AI 📈\n\n"
                "Không còn phải viết những hàm Excel phức tạp thủ công! Trợ lý AI hiện nay có thể giúp bạn xử lý mọi bảng tính dữ liệu lớn:\n"
                "- Tự động viết công thức Excel dựa trên yêu cầu ngôn ngữ tự nhiên.\n"
                "- Nhận diện lỗi sai trong bảng tính và đề xuất cách sửa đổi lập tức.\n"
                "- Tạo biểu đồ trực quan hóa số liệu kinh doanh chỉ trong 10 giây.\n\n"
                "👉 Đọc bài viết hướng dẫn chi tiết tại:\n"
                "https://khoahocai.com.vn/khoa-hoc-ai-trong-van-van-phong/\n\n"
                "---\n"
                "#KhoaHocAI #ExcelAI #Copilot #ChatGPT #ExcelFormulas #TàiChínhAI #SaoViet"
            ),
            "image_prompt": (
                "A sleek dark-mode laptop screen showing a Microsoft Excel sheet with complex data charts (bar charts, line charts) and a glowing virtual assistant sidebar generating clean formulas. "
                "Professional workspace, elegant soft ambient lighting."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc-ai-trong-van-van-phong/"
        },
        # Day 22
        {
            "offset": 21,
            "message": (
                "📽️ THIẾT KẾ SLIDE THUYẾT TRÌNH ĐẸP MẮT TRONG 5 PHÚT BẰNG AI 📽️\n\n"
                "Bạn mất hàng giờ đồng hồ để căn chỉnh bố cục, tìm ảnh minh họa cho slide thuyết trình? Hãy để Gamma AI và ChatGPT làm thay bạn:\n"
                "✨ Nhập chủ đề bài viết và nhận ngay một bộ slide hoàn chỉnh với tiêu đề, nội dung và bố cục phân chia khoa học.\n"
                "✨ Tự động chèn các hình ảnh AI minh họa sinh động phù hợp ngữ cảnh.\n"
                "✨ Dễ dàng chỉnh sửa trực quan nhanh chóng trước khi xuất file PDF/Powerpoint.\n\n"
                "👉 Khám phá thêm lộ trình nâng tầm kỹ năng văn phòng thông minh tại:\n"
                "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/\n\n"
                "---\n"
                "#KhoaHocAI #GammaAI #AIForPresentations #SlideThuyetTrinh #ProductivityHack #ChatGPT #SaoViet"
            ),
            "image_prompt": (
                "A clean desktop screen showing a slide presentation editor. "
                "Beautiful modern slide layouts with futuristic tech graphics and minimal text. "
                "A glowing magic wand icon floating nearby, indicating AI creation. "
                "Clean vector illustration, pastel theme."
            ),
            "link": "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/"
        },
        # Day 23
        {
            "offset": 22,
            "message": (
                "🥊 SO SÁNH: MAKE.COM VS N8N - LỰA CHỌN NÀO CHO DOANH NGHIỆP CỦA BẠN? 🥊\n\n"
                "Cả hai đều là những gã khổng lồ trong lĩnh vực tự động hóa quy trình (iPaaS), nhưng đâu là công cụ phù hợp với bạn?\n"
                "👉 Make.com: Giao diện kéo thả trực quan tuyệt đẹp, cực kỳ thân thiện với người không chuyên (No-code), setup nhanh chóng.\n"
                "👉 n8n.io: Mạnh mẽ cho lập trình viên (Low-code), khả năng tự lưu trữ miễn phí, tiết kiệm tối đa chi phí vận hành ở quy mô lớn.\n\n"
                "👉 Đọc bài viết hướng dẫn sử dụng Make để làm quen với tự động hóa đầu tiên tại:\n"
                "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/\n\n"
                "---\n"
                "#KhoaHocAI #MakeVsN8n #AutomationTools #NoCodeLowCode #WorkflowManagement #CongNgheMoi"
            ),
            "image_prompt": (
                "Two robots facing each other in a friendly wrestling ring. "
                "One robot is bright colored with rounded shapes (representing Make.com). "
                "The other robot is dark metallic with orange wires (representing n8n). "
                "Abstract background, 3D claymation style, playful and clean."
            ),
            "link": "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/"
        },
        # Day 24
        {
            "offset": 23,
            "message": (
                "🎓 NGÀNH TRÍ TUỆ NHÂN TẠO (AI): HỌC GÌ VÀ LÀM GÌ ĐỂ DẪN ĐẦU XU HƯỚNG? 🎓\n\n"
                "Sức hút từ ngành Trí tuệ Nhân tạo là không thể bàn cãi trong kỷ nguyên số hiện nay. Để tham gia vào thị trường đầy tiềm năng này, bạn cần biết:\n"
                "📚 Những kỹ năng trọng yếu cần tích lũy: Toán tối ưu, lập trình Python, xử lý dữ liệu và thiết kế Prompt.\n"
                "💼 Các vị trí công việc khát nhân lực: Kỹ sư Prompt, Chuyên viên phân tích dữ liệu AI, Kỹ sư hệ thống tự động hóa.\n\n"
                "👉 Đọc ngay bài phân tích định hướng chi tiết tại bài viết dưới đây:\n"
                "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/\n\n"
                "---\n"
                "#KhoaHocAI #NganhAI #ViecLamAI #DaoTaoAI #PromptEngineer #LapTrinhPython #CongNghe2026"
            ),
            "image_prompt": (
                "A young data scientist working at a holographic workstation analyzing neural network models and complex code scripts. "
                "Futuristic laboratory, deep shades of blue and gold light, cinematic depth of field."
            ),
            "link": "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/"
        },
        # Day 25
        {
            "offset": 24,
            "message": (
                "🖼️ TOP 5 CÔNG CỤ AI TẠO HÌNH ẢNH NGHỆ THUẬT ĐỈNH CAO NHẤT HIỆN NAY 🖼️\n\n"
                "Hình ảnh đẹp mắt giúp tăng 80% tỷ lệ giữ chân khách hàng trên các kênh mạng xã hội. Hãy tham khảo 5 công cụ tạo ảnh AI hàng đầu:\n"
                "1. Midjourney: Chất ảnh nghệ thuật, chân thực, độ chi tiết siêu cao.\n"
                "2. DALL-E 3 (OpenAI): Khả năng hiểu ngôn ngữ tự nhiên tốt nhất, tích hợp sẵn trong ChatGPT.\n"
                "3. Stable Diffusion: Mã nguồn mở, tùy biến cực mạnh cho dân chuyên nghiệp.\n"
                "4. Adobe Firefly: Tuyệt vời cho thiết kế thương mại, an toàn về bản quyền.\n"
                "5. Canva AI: Tiện lợi, tích hợp sẵn các template thiết kế đa dạng.\n\n"
                "👉 Tìm hiểu thêm kiến thức cơ bản về ứng dụng AI trong đời sống tại:\n"
                "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/\n\n"
                "---\n"
                "#KhoaHocAI #AIImages #Midjourney #DALLE3 #StableDiffusion #SángTạoẢnhAI #SaoViet"
            ),
            "image_prompt": (
                "A modern art gallery exhibition showing futuristic AI-generated masterpieces framed on clean gallery walls. "
                "Art patrons looking at the glowing digital paintings. "
                "Vibrant colors, elegant architectural lines, realistic photo style."
            ),
            "link": "https://khoahocai.com.vn/ai-la-gi-kien-thuc-ve-tri-tue-nhan-tao-va-cach-ung-dung-trong-cuoc-song/"
        },
        # Day 26
        {
            "offset": 25,
            "message": (
                "✉️ TỰ ĐỘNG HÓA EMAIL MARKETING KẾT HỢP AI VÀ MAKE.COM ✉️\n\n"
                "Gửi hàng nghìn email cá nhân hóa cho từng khách hàng mà không cần bấm nút thủ công? Điều này hoàn toàn khả thi bằng việc kết hợp AI và Make:\n"
                "🔸 Nhận diện thông tin khách hàng đăng ký mới từ Form.\n"
                "🔸 Sử dụng ChatGPT viết nội dung email cá nhân hóa phù hợp với hồ sơ khách hàng.\n"
                "🔸 Tự động gửi email qua Gmail/Mailchimp và cập nhật lịch sử tương tác.\n\n"
                "👉 Học ngay cách thiết lập luồng tự động đầu tiên của bạn tại:\n"
                "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/\n\n"
                "---\n"
                "#KhoaHocAI #EmailMarketing #MakeAutomation #ChatGPT #Personalization #NoCode #Productivity"
            ),
            "image_prompt": (
                "A digital visualization representing personalized emails flying out of a tablet screen like glowing paper airplanes, landing directly into different customer profile cards. "
                "Vibrant and clean design, positive corporate vibes, white background."
            ),
            "link": "https://khoahocai.com.vn/huong-dan-su-dung-make-com-tu-dong-hoa-cong-viec-hang-ngay-phan-1/"
        },
        # Day 27
        {
            "offset": 26,
            "message": (
                "🏢 ĐÀO TẠO AI IN-HOUSE: GIẢI PHÁP TỐI ƯU HÓA HIỆU SUẤT CHO DOANH NGHIỆP 🏢\n\n"
                "Nhân sự chính là cốt lõi của chuyển đổi số doanh nghiệp. Trung tâm Sao Việt cung cấp chương trình đào tạo AI In-house thiết kế riêng biệt:\n"
                "✔️ Phân tích trực tiếp các bài toán khó khăn trong vận hành của từng doanh nghiệp.\n"
                "✔️ Hướng dẫn nhân viên ứng dụng AI tự động hóa báo cáo và tối ưu hóa quy trình làm việc.\n"
                "✔️ Tăng tốc năng suất lao động tổng thể và giảm thiểu chi phí vận hành.\n\n"
                "👉 Liên hệ nhận tư vấn chương trình đào tạo doanh nghiệp tại:\n"
                "https://khoahocai.com.vn/khoa-hoc/\n\n"
                "---\n"
                "#KhoaHocAI #DaoTaoDoanhNghiep #AIInHouse #SaoViet #ChuyenDoiSoDoanhNghiep #TốiƯuChiPhí"
            ),
            "image_prompt": (
                "A modern corporate conference room. "
                "An expert trainer explaining AI data flow diagrams on a glass board to a diverse group of corporate employees holding tablets. "
                "Sunlight, clean, highly professional environment."
            ),
            "link": "https://khoahocai.com.vn/khoa-hoc/"
        },
        # Day 28
        {
            "offset": 27,
            "message": (
                "💬 XÂY DỰNG TRỢ LÝ AI CHĂM SÓC KHÁCH HÀNG TỰ ĐỘNG 24/7 💬\n\n"
                "Không còn lo lắng việc trễ nải tin nhắn của khách hàng vào ban đêm hay những ngày lễ! Trợ lý AI có thể giúp bạn:\n"
                "📞 Tự động trả lời thắc mắc của khách hàng về sản phẩm/dịch vụ theo đúng kịch bản doanh nghiệp.\n"
                "📞 Phân loại khách hàng tiềm năng và thu thập thông tin số điện thoại/email chuyển giao cho sale.\n"
                "📞 Tích hợp đa kênh từ Website, Messenger đến Zalo.\n\n"
                "👉 Tìm hiểu ngay cách lập kế hoạch ứng dụng AI hiệu quả cho người mới tại:\n"
                "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/\n\n"
                "---\n"
                "#KhoaHocAI #AIChatbot #CustomerServiceAI #CustomerCare #TựĐộngHóa #ZaloAI #MessengerBot"
            ),
            "image_prompt": (
                "A friendly floating orb-shaped chatbot with a smiling face screen, interacting with chat bubbles showing customer questions about business services. "
                "Soft, clean tech design, modern workspace background."
            ),
            "link": "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/"
        },
        # Day 29
        {
            "offset": 28,
            "message": (
                "🌐 XU HƯỚNG AI AGENTS: TƯƠNG LAI MỚI CỦA CÔNG NGHỆ THÔNG TIN 🌐\n\n"
                "Tại sao các tập đoàn công nghệ hàng đầu thế giới đang đầu tư hàng tỷ USD vào phát triển AI Agent?\n"
                "⚡ Khả năng làm việc đa nhiệm độc lập không mệt mỏi.\n"
                "⚡ Tiết kiệm thời gian lập trình và sửa lỗi hệ thống (nhờ các Agent lập trình thông minh).\n"
                "⚡ Khả năng kết nối, chia sẻ dữ liệu và tự động ra quyết định logic chặt chẽ.\n\n"
                "👉 Đọc bài viết phân tích xu thế công nghệ AI mới nhất tại:\n"
                "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/\n\n"
                "---\n"
                "#KhoaHocAI #AIAgent #TechTrends #FutureTech #SoftwareAutomation #GoogleAntigravity #n8n"
            ),
            "image_prompt": (
                "A global digital network globe with floating holographic screens displaying code graphs, task workflows, and data visualizations. "
                "Bright glowing lines, dark cybernetic background, representation of global automation."
            ),
            "link": "https://khoahocai.com.vn/nganh-tri-tue-nhan-tao-ai-hoc-gi-va-lam-gi-o-nam-2025/"
        },
        # Day 30
        {
            "offset": 29,
            "message": (
                "🎓 HÀNH TRÌNH CHINH PHỤC CÔNG NGHỆ AI CÙNG TRUNG TÂM SAO VIỆT 🎓\n\n"
                "Đã đến lúc bạn viết tiếp câu chuyện thành công của chính mình bằng việc nắm bắt sức mạnh Trí tuệ Nhân tạo.\n"
                "Tại Trung tâm Đào tạo AI Sao Việt, chúng tôi cam kết mang đến những giá trị thực tế nhất cho học viên:\n"
                "🌟 Học từ gốc: Nắm vững tư duy giải quyết vấn đề bằng công nghệ AI.\n"
                "🌟 Học thực chiến: 100% thời gian thực hành thực tế, cầm tay chỉ việc.\n"
                "🌟 Hỗ trợ trọn đời: Đồng hành giải đáp mọi khó khăn kỹ thuật trong và sau khóa học.\n\n"
                "👉 Hãy bắt đầu ngay hôm nay bằng cách đăng ký lộ trình học tại:\n"
                "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/\n\n"
                "---\n"
                "#KhoaHocAI #DaoTaoAISaoViet #SaoViet #AiThucChien #KyNangSo #HocAiSaoViet #ThanhCongCungAI"
            ),
            "image_prompt": (
                "A group of diverse graduates from an AI academy raising their hands holding certificates, smiling confidently. "
                "In the background, a digital neural network graphic blends with warm studio lights. "
                "Celebratory, inspiring, premium corporate photography style."
            ),
            "link": "https://khoahocai.com.vn/lo-trinh-hoc-ai-cho-nguoi-moi-bat-dau/"
        }
    ]
    
    # Fill data rows
    current_row = 2
    for item in posts_data:
        post_date = start_date + timedelta(days=item["offset"])
        date_str = post_date.strftime("%d/%m/%Y")
        
        ws.cell(row=current_row, column=1, value=date_str)
        ws.cell(row=current_row, column=2, value=item["message"])
        ws.cell(row=current_row, column=3, value=item["image_prompt"])
        ws.cell(row=current_row, column=4, value=item["link"])
        ws.cell(row=current_row, column=5, value="Pending")
        ws.cell(row=current_row, column=6, value="")
        ws.cell(row=current_row, column=7, value="")
        
        current_row += 1
        
    # Styles for cells
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    for row in range(2, current_row):
        # Scheduled Date
        cell_date = ws.cell(row=row, column=1)
        cell_date.alignment = Alignment(horizontal="center", vertical="center")
        
        # Message
        cell_msg = ws.cell(row=row, column=2)
        cell_msg.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Image Prompt
        cell_prompt = ws.cell(row=row, column=3)
        cell_prompt.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        # Link
        cell_link = ws.cell(row=row, column=4)
        cell_link.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell_link.font = Font(name="Segoe UI", size=10, color="0000FF", underline="single")
        
        # Status
        cell_status = ws.cell(row=row, column=5)
        cell_status.alignment = Alignment(horizontal="center", vertical="center")
        cell_status.font = Font(name="Segoe UI", size=10, bold=True, color="FF6600")
        cell_status.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Apply borders to all columns in the row
        for col in range(1, 8):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if col not in [4, 5]: # Leave custom font colors alone
                c.font = Font(name="Segoe UI", size=10)

    # Set column widths
    column_widths = {
        "A": 16, # Date
        "B": 50, # Message
        "C": 45, # Image Prompt
        "D": 35, # Article Link
        "E": 12, # Status
        "F": 22, # Post ID
        "G": 20  # Published Time
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Auto-adjust height for rows based on message length (rough estimate)
    ws.row_dimensions[1].height = 28 # Header row
    for r in range(2, current_row):
        msg_len = len(ws.cell(row=r, column=2).value or "")
        prompt_len = len(ws.cell(row=r, column=3).value or "")
        max_len = max(msg_len / 3.5, prompt_len / 3.0)
        ws.row_dimensions[r].height = max(55, min(220, max_len))
        
    # Save file
    file_name = "Facebook_Post_Schedule.xlsx"
    wb.save(file_name)
    print(f"Content plan generated successfully and saved to {os.path.abspath(file_name)}")
    return os.path.abspath(file_name)

if __name__ == "__main__":
    create_excel_schedule()
