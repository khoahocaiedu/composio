import os
import sys
import shutil
import time
import json
from datetime import datetime
from dotenv import load_dotenv
load_dotenv()

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import PatternFill, Font

from composio import Composio

def safe_print(text):
    try:
        print(text)
    except UnicodeEncodeError:
        print(str(text).encode('ascii', errors='replace').decode('ascii'))

def run_post_from_excel(image_path):
    excel_path = "Facebook_Post_Schedule.xlsx"
    if not os.path.exists(excel_path):
        safe_print(f"Error: {excel_path} not found.")
        sys.exit(1)
        
    safe_print(f"Loading {excel_path}...")
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Find the first pending post
    target_row = None
    for r in range(2, ws.max_row + 1):
        status_val = ws.cell(row=r, column=5).value
        if status_val == "Pending":
            target_row = r
            break
            
    if target_row is None:
        safe_print("No 'Pending' posts found in the schedule.")
        wb.close()
        return
        
    # Extract data
    scheduled_date = ws.cell(row=target_row, column=1).value
    message = ws.cell(row=target_row, column=2).value
    image_prompt = ws.cell(row=target_row, column=3).value
    article_link = ws.cell(row=target_row, column=4).value
    
    safe_print(f"\n--- Found Pending Post for Row {target_row} ---")
    safe_print(f"Scheduled Date: {scheduled_date}")
    safe_print(f"Link: {article_link}")
    safe_print(f"Message preview:\n{message[:150]}...")
    safe_print(f"Image Prompt preview:\n{image_prompt[:150]}...")
    safe_print("------------------------------------\n")
    
    # Verify image path
    if not os.path.exists(image_path):
        safe_print(f"Error: Image not found at {image_path}")
        wb.close()
        sys.exit(1)
        
    # Copy image to Composio temp folder
    temp_dir = os.path.expanduser("~/.composio/temp")
    os.makedirs(temp_dir, exist_ok=True)
    target_photo = os.path.join(temp_dir, "temp_post_photo.png")
    safe_print(f"Copying photo to temp directory: {target_photo}")
    shutil.copy2(image_path, target_photo)
    
    # Initialize Composio
    safe_print("Initializing Composio and executing FACEBOOK_CREATE_PHOTO_POST...")
    composio = Composio(dangerously_allow_auto_upload_download_files=True)
    user_id = "pg-test-f88a0cbe-fcae-46a0-b516-4204597f4607"
    page_id = "786741061182982"
    
    try:
        post_result = composio.tools.execute(
            slug="FACEBOOK_CREATE_PHOTO_POST",
            user_id=user_id,
            arguments={
                "page_id": page_id,
                "photo": target_photo,
                "message": message
            },
            dangerously_skip_version_check=True
        )
        
        # Check result and extract post ID
        safe_print("Response received from Composio API.")
        post_id = None
        if isinstance(post_result, dict):
            post_id = post_result.get("id") or post_result.get("post_id") or post_result.get("data", {}).get("id")
        
        if not post_id:
            # Try parsing raw string
            res_str = str(post_result)
            safe_print(f"Raw Result: {res_str}")
            # Try standard extraction
            import re
            m = re.search(r"'(?:id|post_id)':\s*'([^']+)'", res_str)
            if m:
                post_id = m.group(1)
            else:
                post_id = "Unknown_ID"
                
        safe_print(f"Published successfully! Post ID: {post_id}")
        
        # Update Excel row
        current_time_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        
        # Status Column
        ws.cell(row=target_row, column=5, value="Success")
        ws.cell(row=target_row, column=5).font = Font(name="Segoe UI", size=10, bold=True, color="008000")
        ws.cell(row=target_row, column=5).fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
        
        # Post ID Column
        ws.cell(row=target_row, column=6, value=post_id)
        
        # Published Time Column
        ws.cell(row=target_row, column=7, value=current_time_str)
        
        # Save Excel file
        wb.save(excel_path)
        safe_print(f"Excel updated successfully for row {target_row}.")
        
    except Exception as e:
        safe_print(f"Error executing post: {e}")
        # Mark as Failed in Excel
        ws.cell(row=target_row, column=5, value="Failed")
        ws.cell(row=target_row, column=5).font = Font(name="Segoe UI", size=10, bold=True, color="FF0000")
        ws.cell(row=target_row, column=5).fill = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")
        wb.save(excel_path)
        safe_print("Excel updated with Failed status.")
        sys.exit(1)
    finally:
        wb.close()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        safe_print("Usage: python run_excel_post.py <path_to_image>")
        sys.exit(1)
    run_post_from_excel(sys.argv[1])
